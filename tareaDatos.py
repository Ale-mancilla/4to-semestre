import string
import openpyxl
from pathlib import Path 
import math
# 1. Genera una lista de 100 elementos con multiplos de 3
multiplos = list(range(3,303,3))

datosCTfile = Path('datosCT.xlsx') 
datosPEfile = Path('datosPE.xlsx')
datosSheet = openpyxl.load_workbook(datosCTfile).active 
datosPEsheet = openpyxl.load_workbook(datosPEfile).active

# 2. Calcula el indice de masa corporal de cada persona (datosPE.xlsx)
imc = lambda peso, estatura: round(peso / (estatura**2), 1)
indices = [] 

# Calcular limites de la hoja de calculo 
for row in datosPEsheet.iter_rows(2,101):
    indice = imc(float(row[1].value),float(row[2].value))
    indices.append(indice)  
print("IMC's:",indices)
# Indices almacenados en: indices[]  

# 3.1 Calcula la utilidad (ingreso-gasto) 
# 3.2 Encuentra el promedio de Gastos en los meses que la utilidad fue negativa 
# 3.3 Encuentra los ingresos totales del año 

# Construir un diccionario global 
class Month:
    def __init__(self, name: string, ingreso: int, gasto: int) -> None:
        self.name = name
        self.ingreso = ingreso
        self.gasto = gasto
        self.utilidad = ingreso - gasto
        pass 
ingresos = 0
negativos = 0
utilidad = 0
for row in datosSheet.iter_rows(2,13):
    mes = Month(row[0].value, int(row[1].value), int(row[2].value))
    ingresos += mes.ingreso 
    utilidad += mes.utilidad
    if mes.utilidad < 0:
        negativos += mes.gasto

print("Ingresos:",ingresos) 
print("Utilidad", utilidad)
print("Promedio de gasto de mesescon utilidad negativo:", negativos / 12)
